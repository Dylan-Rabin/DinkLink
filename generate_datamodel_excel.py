#!/usr/bin/env python3
"""
Generates DinkLink-DataModel.xlsx — one sheet per table.
Option A + MVP: user_progression merged into user_profiles; level_tiers removed.
MVP additions: streak fields, challenge flags, shot_events table, corrected XP sources.
Run: python3 generate_datamodel_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_DARK   = "1F2D3D"   # dark navy  – sheet header row
CLR_HEADER_MID    = "2E7D9E"   # teal       – column-header row
CLR_HEADER_LIGHT  = "D6EAF8"   # pale blue  – alternating row tint
CLR_PK            = "FFF3CD"   # amber      – primary key highlight
CLR_FK            = "D5F5E3"   # mint green – foreign key highlight
CLR_NEW           = "EDE7F6"   # lavender   – new tables
CLR_EXISTING      = "E8F5E9"   # pale green – existing tables
CLR_WHITE         = "FFFFFF"
CLR_BORDER        = "B0BEC5"

FONT_TITLE  = Font(name="Calibri", bold=True, size=13, color=CLR_WHITE)
FONT_COL_H  = Font(name="Calibri", bold=True, size=10, color=CLR_WHITE)
FONT_BODY   = Font(name="Calibri", size=10)
FONT_BODY_B = Font(name="Calibri", bold=True, size=10)
FONT_SMALL  = Font(name="Calibri", size=9, color="555555")

thin = Side(style="thin",  color=CLR_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── Table definitions ─────────────────────────────────────────────────────────
# Each table: { name, status, description, columns: [{...}] }
# column keys: name, type, constraints, description, lookup
# constraints: PK | FK | UNIQUE | NOT NULL | DEFAULT x
# lookup: string shown in the "Linked To" cell

TABLES = [
    # ── LOCAL / SwiftData ────────────────────────────────────────────────────
    {
        "name": "PlayerProfile",
        "status": "EXISTING — Local (SwiftData)",
        "layer": "local",
        "description": (
            "On-device SwiftData model. One record per device user. "
            "Synced remotely to user_profiles after Phase 1 cutover."
        ),
        "columns": [
            {"name": "id",                   "type": "UUID",    "constraints": "PK · UNIQUE",    "description": "Auto-generated UUID, primary key.",                                     "lookup": ""},
            {"name": "name",                 "type": "String",  "constraints": "NOT NULL",        "description": "Player's display name.",                                                "lookup": ""},
            {"name": "locationName",         "type": "String",  "constraints": "",                "description": "Freetext home location. Superseded by SavedLocation.isHome after Phase 3.", "lookup": "SavedLocation.isHome = true → placeName"},
            {"name": "dominantArmRawValue",  "type": "String",  "constraints": "NOT NULL",        "description": "Raw value of DominantArm enum. Allowed: Right | Left | Ambidextrous.",  "lookup": "Enum: DominantArm"},
            {"name": "skillLevelRawValue",   "type": "String",  "constraints": "NOT NULL",        "description": "Raw value of SkillLevel enum. Allowed: Beginner | Intermediate | Advanced | Tournament.", "lookup": "Enum: SkillLevel"},
            {"name": "syncedPaddleName",     "type": "String",  "constraints": "",                "description": "Name of the last BLE-connected paddle device.",                          "lookup": ""},
            {"name": "completedOnboarding",  "type": "Bool",    "constraints": "DEFAULT false",   "description": "True once the user has finished onboarding flow.",                       "lookup": ""},
            {"name": "supabaseProfileSynced","type": "Bool",    "constraints": "DEFAULT false",   "description": "NEW (Phase 1). True once this profile has been pushed to Supabase.",     "lookup": ""},
            {"name": "gpnUsername",          "type": "String?", "constraints": "",                "description": "NEW (Phase 6). User's Global Pickleball Network username handle. Used with GPN OAuth API.",       "lookup": "user_profiles.gpn_username"},
            {"name": "homeLocationLabel",    "type": "String?", "constraints": "",                "description": "NEW (Phase 3). Convenience cache of the home location label.",           "lookup": "SavedLocation.label WHERE isHome = true"},
            {"name": "currentStreak",        "type": "Int",    "constraints": "DEFAULT 0",      "description": "NEW (MVP). Consecutive daily-play streak. Incremented when lastActiveDate was yesterday.", "lookup": "user_profiles.current_streak"},
            {"name": "longestStreak",        "type": "Int",    "constraints": "DEFAULT 0",      "description": "NEW (MVP). All-time best consecutive daily-play streak.",                 "lookup": "user_profiles.longest_streak"},
            {"name": "lastActiveDate",       "type": "Date?",  "constraints": "",               "description": "NEW (MVP). Last calendar day a session was completed. Used for streak calculation.", "lookup": ""},
        ],
    },
    {
        "name": "StoredGameSession",
        "status": "EXISTING — Local (SwiftData)",
        "layer": "local",
        "description": (
            "On-device SwiftData model. One record per completed game session. "
            "Replicated to remote game_sessions table after Phase 2 cutover."
        ),
        "columns": [
            {"name": "id",                   "type": "UUID",    "constraints": "PK · UNIQUE",   "description": "Auto-generated UUID, primary key. Also used as remote game_sessions.id.", "lookup": "game_sessions.id"},
            {"name": "modeRawValue",         "type": "String",  "constraints": "NOT NULL",       "description": "Raw value of GameMode enum. Allowed: Dink Sinks | Volley Wallies | The Real Deal | Pickle Cup.", "lookup": "Enum: GameMode"},
            {"name": "startDate",            "type": "Date",    "constraints": "NOT NULL",       "description": "UTC timestamp when the session started.",                                "lookup": ""},
            {"name": "endDate",              "type": "Date",    "constraints": "NOT NULL",       "description": "UTC timestamp when the session ended. Used for sort order.",             "lookup": ""},
            {"name": "playerOneName",        "type": "String",  "constraints": "NOT NULL",       "description": "Display name of player one.",                                            "lookup": ""},
            {"name": "playerTwoName",        "type": "String",  "constraints": "NOT NULL",       "description": "Display name of player two. 'Solo' if single-player mode.",             "lookup": ""},
            {"name": "playerOneScore",       "type": "Int",     "constraints": "DEFAULT 0",      "description": "Final score for player one.",                                            "lookup": ""},
            {"name": "playerTwoScore",       "type": "Int",     "constraints": "DEFAULT 0",      "description": "Final score for player two.",                                            "lookup": ""},
            {"name": "averageSwingSpeed",    "type": "Double",  "constraints": "DEFAULT 0",      "description": "Mean paddle swing speed across all hits in the session (MPH).",         "lookup": ""},
            {"name": "maxSwingSpeed",        "type": "Double",  "constraints": "DEFAULT 0",      "description": "Maximum recorded swing speed in the session (MPH). Used for PB tracking.", "lookup": ""},
            {"name": "sweetSpotPercentage",  "type": "Double",  "constraints": "DEFAULT 0",      "description": "Percentage of hits landing in the paddle sweet spot (0–100).",          "lookup": ""},
            {"name": "totalHits",            "type": "Int",     "constraints": "DEFAULT 0",      "description": "Total number of shot events recorded.",                                  "lookup": ""},
            {"name": "winnerName",           "type": "String",  "constraints": "NOT NULL",       "description": "Display name of the session winner.",                                    "lookup": ""},
            {"name": "longestStreak",        "type": "Int",     "constraints": "DEFAULT 0",      "description": "Best consecutive dink streak achieved (Dink Sinks mode).",              "lookup": ""},
            {"name": "totalValidVolleys",    "type": "Int",     "constraints": "DEFAULT 0",      "description": "Total clean volleys ≥ 15 MPH (Volley Wallies mode).",                  "lookup": ""},
            {"name": "bestRallyLength",      "type": "Int",     "constraints": "DEFAULT 0",      "description": "Longest rally measured in total hits (The Real Deal mode).",            "lookup": ""},
            {"name": "remoteID",             "type": "UUID?",   "constraints": "",               "description": "NEW (Phase 2). Populated once this session is confirmed in Supabase.",  "lookup": "game_sessions.id"},
            {"name": "isDirty",              "type": "Bool",    "constraints": "DEFAULT true",   "description": "NEW (Phase 2). True = pending sync to Supabase.",                       "lookup": ""},
            {"name": "isChallenge",          "type": "Bool",    "constraints": "DEFAULT false",  "description": "NEW (MVP). True = this was a challenge match. If user wins, awards +100 XP.", "lookup": "game_sessions.is_challenge"},
            {"name": "isPickleCupWin",       "type": "Bool",    "constraints": "DEFAULT false",  "description": "NEW (MVP). True = user won the full Pickle Cup sequence.",               "lookup": "game_sessions.is_pickle_cup_win"},
        ],
    },
    {
        "name": "SavedLocation",
        "status": "NEW — Local (SwiftData)",
        "layer": "local",
        "description": (
            "NEW (Phase 3). On-device SwiftData model. "
            "Replaces the single PlayerProfile.locationName string with "
            "support for multiple named locations per user. Synced to saved_locations."
        ),
        "columns": [
            {"name": "id",          "type": "UUID",    "constraints": "PK · UNIQUE",   "description": "Auto-generated UUID, primary key.",                                           "lookup": "saved_locations.id"},
            {"name": "label",       "type": "String",  "constraints": "NOT NULL",       "description": "User-facing name for the location (e.g. 'Home', 'Rec Center').",            "lookup": ""},
            {"name": "placeName",   "type": "String",  "constraints": "NOT NULL",       "description": "Resolved venue or area name (e.g. 'Lincoln Park Courts').",               "lookup": ""},
            {"name": "address",     "type": "String?", "constraints": "",               "description": "Optional full street address.",                                               "lookup": ""},
            {"name": "latitude",    "type": "Double?", "constraints": "",               "description": "Decimal latitude for map/weather features.",                                  "lookup": ""},
            {"name": "longitude",   "type": "Double?", "constraints": "",               "description": "Decimal longitude for map/weather features.",                                 "lookup": ""},
            {"name": "isHome",      "type": "Bool",    "constraints": "DEFAULT false",  "description": "True if this is the user's primary/home location. At most one per user.",   "lookup": "PlayerProfile.locationName (legacy)"},
            {"name": "supabaseID",  "type": "UUID?",   "constraints": "",               "description": "Remote saved_locations.id once synced.",                                     "lookup": "saved_locations.id"},
            {"name": "isDirty",     "type": "Bool",    "constraints": "DEFAULT true",   "description": "True = pending sync to Supabase.",                                           "lookup": ""},
            {"name": "createdAt",   "type": "Date",    "constraints": "NOT NULL",       "description": "Timestamp when this location was created locally.",                          "lookup": ""},
        ],
    },
    {
        "name": "SyncQueueItem",
        "status": "NEW — Local (SwiftData)",
        "layer": "local",
        "description": (
            "NEW (Phase 1). Offline write queue. When the device has no network, "
            "any Supabase-bound write is serialised here. SyncService drains this "
            "table in chronological order when connectivity is restored."
        ),
        "columns": [
            {"name": "id",          "type": "UUID",   "constraints": "PK · UNIQUE",   "description": "Auto-generated UUID.",                                                                 "lookup": ""},
            {"name": "operation",   "type": "String", "constraints": "NOT NULL",       "description": "Type of operation. Allowed: upsert_profile | save_session | upsert_location | award_badge | xp_events. NOTE: upsert_progression removed (XP now on user_profiles).", "lookup": ""},
            {"name": "tableName",   "type": "String", "constraints": "NOT NULL",       "description": "Supabase table this item targets (e.g. 'game_sessions').",                           "lookup": ""},
            {"name": "payload",     "type": "Data",   "constraints": "NOT NULL",       "description": "JSON-encoded request body to be replayed against the Supabase REST endpoint.",       "lookup": ""},
            {"name": "createdAt",   "type": "Date",   "constraints": "NOT NULL",       "description": "Timestamp of enqueue. Items are drained oldest-first.",                              "lookup": ""},
            {"name": "retryCount",  "type": "Int",    "constraints": "DEFAULT 0",      "description": "Number of failed drain attempts. Items with retryCount ≥ 5 are abandoned.",         "lookup": ""},
        ],
    },

    # ── REMOTE / Supabase ────────────────────────────────────────────────────
    {
        "name": "auth.users",
        "status": "EXISTING — Remote (Supabase built-in)",
        "layer": "remote",
        "description": (
            "Supabase-managed authentication table. Not customised directly. "
            "All custom profile data extends into user_profiles via a 1:1 FK."
        ),
        "columns": [
            {"name": "id",    "type": "uuid",  "constraints": "PK",         "description": "Supabase-generated user UUID. Referenced as FK by all user-owned tables.", "lookup": "user_profiles.user_id · saved_locations.user_id · user_badges.user_id · game_sessions.user_id · xp_events.user_id · comments.user_id · comment_likes.user_id"},
            {"name": "email", "type": "text",  "constraints": "UNIQUE",     "description": "User's email address used for authentication.",                              "lookup": ""},
        ],
    },
    {
        "name": "user_profiles",
        "status": "NEW — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "NEW (Phase 1). Public-facing profile extending auth.users. "
            "OPTION A: total_xp and xp_updated_at merged in from removed user_progression table. "
            "Level/rank resolved in Swift via ProgressionService constants (no level_tiers DB table). "
            "MVP: streak fields added (current_streak, longest_streak, last_active_date). "
            "GPN integration via real OAuth API at globalpickleball.network/developers. "
            "Row Level Security: users can only read/write their own row."
        ),
        "columns": [
            {"name": "user_id",            "type": "uuid",          "constraints": "PK · FK",          "description": "References auth.users(id). Deletes cascade.",                              "lookup": "auth.users.id"},
            {"name": "display_name",       "type": "text",          "constraints": "NOT NULL DEFAULT ''","description": "User's chosen display name, synced from PlayerProfile.name.",             "lookup": ""},
            {"name": "avatar_url",         "type": "text",          "constraints": "nullable",          "description": "URL to the user's profile avatar image.",                                "lookup": ""},
            {"name": "total_xp",           "type": "int",           "constraints": "NOT NULL DEFAULT 0", "description": "MERGED (Option A). Cumulative XP. Was user_progression.total_xp. Level resolved in Swift using ProgressionService.levelThresholds.", "lookup": "ProgressionService.levelThresholds (Swift)"},
            {"name": "xp_updated_at",      "type": "timestamptz",   "constraints": "nullable",          "description": "MERGED (Option A). Timestamp of last XP update. Was user_progression.updated_at.", "lookup": ""},
            {"name": "gpn_username",       "type": "text",          "constraints": "nullable",          "description": "User's username handle on Global Pickleball Network.",                    "lookup": "GPN: globalpickleball.network/profile/{gpn_username}"},
            {"name": "gpn_profile_url",    "type": "text",          "constraints": "nullable",          "description": "Full cached URL to the user's GPN profile page.",                        "lookup": ""},
            {"name": "gpn_singles_level",  "type": "numeric(4,2)",  "constraints": "nullable",          "description": "GPN-calculated singles playing level (e.g. 3.50). Synced by Edge Function.", "lookup": ""},
            {"name": "gpn_doubles_level",  "type": "numeric(4,2)",  "constraints": "nullable",          "description": "GPN-calculated doubles playing level. Synced by Edge Function.",          "lookup": ""},
            {"name": "gpn_dupr_rating",    "type": "numeric(4,2)",  "constraints": "nullable",          "description": "DUPR rating surfaced via GPN's DUPR partnership. Synced by Edge Function.","lookup": "DUPR: dupr.com"},
            {"name": "gpn_last_synced_at", "type": "timestamptz",   "constraints": "nullable",          "description": "Timestamp of the most recent successful GPN data sync via OAuth API.",   "lookup": ""},
            {"name": "current_streak",     "type": "int",          "constraints": "NOT NULL DEFAULT 0", "description": "NEW (MVP). Consecutive daily-play streak. Used for 5-day streak XP award (+75 XP).", "lookup": "PlayerProfile.currentStreak"},
            {"name": "longest_streak",     "type": "int",          "constraints": "NOT NULL DEFAULT 0", "description": "NEW (MVP). All-time best consecutive daily-play streak.",               "lookup": "PlayerProfile.longestStreak"},
            {"name": "last_active_date",   "type": "date",         "constraints": "nullable",           "description": "NEW (MVP). Last calendar day with a completed session. Used for streak computation.", "lookup": ""},
            {"name": "created_at",         "type": "timestamptz",  "constraints": "NOT NULL DEFAULT now()", "description": "Row creation timestamp.",                                           "lookup": ""},
            {"name": "updated_at",         "type": "timestamptz",  "constraints": "NOT NULL DEFAULT now()", "description": "Last update timestamp. Updated via trigger on any column change.",   "lookup": ""},
        ],
    },
    {
        "name": "game_sessions",
        "status": "NEW — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "NEW (Phase 2). Remote replica of StoredGameSession. "
            "Enables multi-device access and cloud backup. "
            "Written by SyncService drain or directly when online. "
            "MVP: is_challenge and is_pickle_cup_win flags added for XP awarding. "
            "Row Level Security: users can only read/write their own sessions."
        ),
        "columns": [
            {"name": "id",                    "type": "uuid",        "constraints": "PK",                 "description": "Same UUID as StoredGameSession.id on device.",                          "lookup": "StoredGameSession.id"},
            {"name": "user_id",               "type": "uuid",        "constraints": "FK · NOT NULL",      "description": "Owner of this session.",                                                "lookup": "user_profiles.user_id"},
            {"name": "mode",                  "type": "text",        "constraints": "NOT NULL",           "description": "Game mode. Allowed: Dink Sinks | Volley Wallies | The Real Deal | Pickle Cup.", "lookup": "Enum: GameMode"},
            {"name": "start_date",            "type": "timestamptz", "constraints": "NOT NULL",           "description": "Session start time (UTC).",                                             "lookup": ""},
            {"name": "end_date",              "type": "timestamptz", "constraints": "NOT NULL",           "description": "Session end time (UTC). Used for ordering.",                            "lookup": ""},
            {"name": "player_one_name",       "type": "text",        "constraints": "NOT NULL",           "description": "Display name of player one.",                                           "lookup": ""},
            {"name": "player_two_name",       "type": "text",        "constraints": "NOT NULL",           "description": "Display name of player two. 'Solo' if single-player.",                 "lookup": ""},
            {"name": "player_one_score",      "type": "int",         "constraints": "NOT NULL DEFAULT 0", "description": "Final score for player one.",                                           "lookup": ""},
            {"name": "player_two_score",      "type": "int",         "constraints": "NOT NULL DEFAULT 0", "description": "Final score for player two.",                                           "lookup": ""},
            {"name": "average_swing_speed",   "type": "float8",      "constraints": "NOT NULL DEFAULT 0", "description": "Mean swing speed across all hits (MPH).",                              "lookup": ""},
            {"name": "max_swing_speed",       "type": "float8",      "constraints": "NOT NULL DEFAULT 0", "description": "Maximum swing speed recorded (MPH). Used for personal best tracking.", "lookup": ""},
            {"name": "sweet_spot_percentage", "type": "float8",      "constraints": "NOT NULL DEFAULT 0", "description": "Percentage of sweet-spot hits (0–100).",                              "lookup": ""},
            {"name": "total_hits",            "type": "int",         "constraints": "NOT NULL DEFAULT 0", "description": "Total shot events in the session.",                                    "lookup": ""},
            {"name": "winner_name",           "type": "text",        "constraints": "NOT NULL DEFAULT ''","description": "Display name of the session winner.",                                   "lookup": ""},
            {"name": "longest_streak",        "type": "int",         "constraints": "NOT NULL DEFAULT 0", "description": "Best consecutive dink streak (Dink Sinks mode).",                    "lookup": ""},
            {"name": "total_valid_volleys",   "type": "int",         "constraints": "NOT NULL DEFAULT 0", "description": "Total clean volleys ≥ 15 MPH (Volley Wallies mode).",                "lookup": ""},
            {"name": "best_rally_length",     "type": "int",         "constraints": "NOT NULL DEFAULT 0", "description": "Longest rally in hits (The Real Deal mode).",                        "lookup": ""},
            {"name": "is_challenge",          "type": "bool",        "constraints": "NOT NULL DEFAULT false", "description": "NEW (MVP). Challenge match flag. Winner earns +100 XP.",              "lookup": "StoredGameSession.isChallenge"},
            {"name": "is_pickle_cup_win",    "type": "bool",        "constraints": "NOT NULL DEFAULT false", "description": "NEW (MVP). True = user won the full Pickle Cup sequence.",           "lookup": "StoredGameSession.isPickleCupWin"},
            {"name": "created_at",            "type": "timestamptz", "constraints": "NOT NULL DEFAULT now()", "description": "Row creation timestamp.",                                         "lookup": ""},
        ],
    },
    {
        "name": "shot_events",
        "status": "NEW — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "NEW (Phase 2). Remote record of individual shot events per session. "
            "Per MVP spec section 6, ShotEvent includes gameId and sessionId as persistent fields. "
            "Enables future biomechanics analysis and shot-level replays. "
            "Row Level Security: users can only read/write their own rows."
        ),
        "columns": [
            {"name": "id",             "type": "uuid",        "constraints": "PK DEFAULT gen_random_uuid()", "description": "Auto-generated UUID.",                                                    "lookup": ""},
            {"name": "session_id",     "type": "uuid",        "constraints": "FK · NOT NULL",      "description": "The game session this shot belongs to. Cascade deletes.",                       "lookup": "game_sessions.id"},
            {"name": "user_id",        "type": "uuid",        "constraints": "FK · NOT NULL",      "description": "Shot owner.",                                                                   "lookup": "user_profiles.user_id"},
            {"name": "timestamp",      "type": "timestamptz", "constraints": "NOT NULL",           "description": "Exact UTC time of the paddle hit event from BLE.",                            "lookup": ""},
            {"name": "speed_mph",      "type": "float8",      "constraints": "NOT NULL",           "description": "Estimated swing speed from BLE estimatedSwingSpeed field (MPH).",               "lookup": ""},
            {"name": "hit_sweet_spot", "type": "bool",        "constraints": "NOT NULL",           "description": "True if paddle BLE event flagged sweetSpotHit.",                              "lookup": ""},
            {"name": "spin_rpm",       "type": "float8",      "constraints": "NOT NULL DEFAULT 0", "description": "Spin rate if available from paddle firmware.",                               "lookup": ""},
            {"name": "created_at",     "type": "timestamptz", "constraints": "NOT NULL DEFAULT now()", "description": "Row creation timestamp.",                                               "lookup": ""},
        ],
    },
    {
        "name": "saved_locations",
        "status": "NEW — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "NEW (Phase 3). Remote replica of SavedLocation SwiftData model. "
            "Supports multiple named locations per user. Partial unique index "
            "ensures each user has at most one home location. "
            "Row Level Security: users can only read/write their own rows."
        ),
        "columns": [
            {"name": "id",         "type": "uuid",      "constraints": "PK",                      "description": "Same UUID as SavedLocation.supabaseID on device.",                          "lookup": "SavedLocation.supabaseID"},
            {"name": "user_id",    "type": "uuid",      "constraints": "FK · NOT NULL",           "description": "Owner of this location.",                                                   "lookup": "user_profiles.user_id"},
            {"name": "label",      "type": "text",      "constraints": "NOT NULL DEFAULT 'Home'", "description": "User-facing name (e.g. 'Home', 'Rec Center', 'Work').",                   "lookup": ""},
            {"name": "place_name", "type": "text",      "constraints": "NOT NULL",                "description": "Resolved venue or area name.",                                              "lookup": ""},
            {"name": "address",    "type": "text",      "constraints": "nullable",                "description": "Optional full street address.",                                             "lookup": ""},
            {"name": "latitude",   "type": "float8",    "constraints": "nullable",                "description": "Decimal latitude. Used with longitude for map/weather features.",          "lookup": ""},
            {"name": "longitude",  "type": "float8",    "constraints": "nullable",                "description": "Decimal longitude.",                                                        "lookup": ""},
            {"name": "is_home",    "type": "bool",      "constraints": "NOT NULL DEFAULT false",  "description": "True = primary home location. PARTIAL UNIQUE INDEX on (user_id) WHERE is_home = true.", "lookup": "PlayerProfile.locationName (legacy migration)"},
            {"name": "created_at", "type": "timestamptz","constraints": "NOT NULL DEFAULT now()", "description": "Row creation timestamp.",                                                   "lookup": ""},
        ],
    },
    {
        "name": "xp_events",
        "status": "EXISTING — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "Existing Supabase table. Immutable event log of every XP award. "
            "Each session produces multiple rows (one per XPBreakdownItem). "
            "Used for auditing and future leaderboard features. "
            "Row Level Security: users can only read/write their own rows."
        ),
        "columns": [
            {"name": "id",       "type": "uuid",   "constraints": "PK DEFAULT gen_random_uuid()", "description": "Auto-generated UUID.",                                                   "lookup": ""},
            {"name": "user_id",  "type": "uuid",   "constraints": "FK · NOT NULL",               "description": "References auth.users(id).",                                             "lookup": "auth.users.id"},
            {"name": "source",   "type": "text",   "constraints": "NOT NULL",                    "description": "XP source label. MVP values: 'Complete session' +50, '10+ clean hits' +20, 'Personal best' +40, 'Played with a friend' +30, '5-day streak' +75, 'Challenge win' +100.", "lookup": ""},
            {"name": "xp",       "type": "int",    "constraints": "NOT NULL",                    "description": "XP amount awarded for this event.",                                      "lookup": ""},
            {"name": "metadata", "type": "jsonb",  "constraints": "nullable",                    "description": "Arbitrary key-value context. Keys include: sync_type, session_count, previous_remote_xp.", "lookup": ""},
            {"name": "created_at","type":"timestamptz","constraints": "NOT NULL DEFAULT now()",   "description": "Event timestamp. XP events must be applied in created_at ASC order.",   "lookup": ""},
        ],
    },
    {
        "name": "badges",
        "status": "NEW — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "NEW (Phase 4). Catalog of all available badges and awards. "
            "Server-side source of truth — the app fetches this list to show "
            "locked/unlocked state without any hardcoded badge logic in Swift. "
            "Public read; admin-only write."
        ),
        "columns": [
            {"name": "id",          "type": "uuid",      "constraints": "PK DEFAULT gen_random_uuid()", "description": "UUID primary key.",                                                  "lookup": "user_badges.badge_id"},
            {"name": "key",         "type": "text",      "constraints": "UNIQUE · NOT NULL",    "description": "Machine-readable badge code. E.g. 'first_session', '100_hits', 'gold_tier'. Used in app logic.", "lookup": ""},
            {"name": "name",        "type": "text",      "constraints": "NOT NULL",             "description": "Display name shown in the UI. E.g. 'First Serve', 'Centurion'.",          "lookup": ""},
            {"name": "description", "type": "text",      "constraints": "NOT NULL DEFAULT ''",  "description": "Tooltip/detail text explaining how to earn the badge.",                    "lookup": ""},
            {"name": "icon_url",    "type": "text",      "constraints": "NOT NULL DEFAULT ''",  "description": "URL to the badge image asset.",                                            "lookup": ""},
            {"name": "badge_type",  "type": "text",      "constraints": "NOT NULL CHECK",       "description": "Category of badge. Allowed: achievement | milestone | tournament | level.", "lookup": "Enum: BadgeType"},
            {"name": "xp_reward",   "type": "int",       "constraints": "NOT NULL DEFAULT 0",   "description": "Bonus XP granted when this badge is first earned.",                       "lookup": "xp_events.xp"},
            {"name": "is_hidden",   "type": "bool",      "constraints": "NOT NULL DEFAULT false","description": "True = secret badge. Hidden in locked state to create surprise reveals.", "lookup": ""},
            {"name": "created_at",  "type": "timestamptz","constraints": "NOT NULL DEFAULT now()","description": "Row creation timestamp.",                                               "lookup": ""},
        ],
    },
    {
        "name": "user_badges",
        "status": "NEW — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "NEW (Phase 4). Junction table linking users to earned badges. "
            "UNIQUE(user_id, badge_id) prevents duplicate awards. "
            "session_id optionally traces which session triggered the award. "
            "Row Level Security: users can only read/write their own rows."
        ),
        "columns": [
            {"name": "id",         "type": "uuid",      "constraints": "PK DEFAULT gen_random_uuid()", "description": "UUID primary key.",                                              "lookup": ""},
            {"name": "user_id",    "type": "uuid",      "constraints": "FK · NOT NULL",      "description": "Badge recipient.",                                                       "lookup": "user_profiles.user_id"},
            {"name": "badge_id",   "type": "uuid",      "constraints": "FK · NOT NULL",      "description": "The badge that was earned.",                                             "lookup": "badges.id"},
            {"name": "awarded_at", "type": "timestamptz","constraints": "NOT NULL DEFAULT now()","description": "Timestamp when the badge was awarded.",                             "lookup": ""},
            {"name": "source",     "type": "text",      "constraints": "nullable",           "description": "Human-readable reason. E.g. 'Completed first session on 2026-04-13'.",  "lookup": ""},
            {"name": "session_id", "type": "uuid",      "constraints": "nullable",           "description": "Optional reference to the game session that triggered this badge.",     "lookup": "game_sessions.id"},
            {"name": "UNIQUE",     "type": "—",         "constraints": "(user_id, badge_id)","description": "Constraint ensuring each badge is awarded at most once per user.",       "lookup": ""},
        ],
    },
    {
        "name": "comments",
        "status": "EXISTING — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "Existing Supabase table. Public comments on game sessions or "
            "other feed items. item_id is a generic UUID reference (not a "
            "strict FK) to support multiple commentable entity types."
        ),
        "columns": [
            {"name": "id",          "type": "uuid",       "constraints": "PK",                    "description": "UUID primary key.",                                                      "lookup": "comment_likes.comment_id"},
            {"name": "item_id",     "type": "uuid",       "constraints": "NOT NULL",              "description": "Generic reference to the item being commented on (e.g. a game session).", "lookup": "game_sessions.id (soft ref)"},
            {"name": "user_id",     "type": "uuid",       "constraints": "FK · NOT NULL",         "description": "Author of the comment.",                                                 "lookup": "auth.users.id"},
            {"name": "author_name", "type": "text",       "constraints": "NOT NULL",              "description": "Display name of the comment author at time of posting.",                 "lookup": ""},
            {"name": "body",        "type": "text",       "constraints": "NOT NULL",              "description": "Comment text content.",                                                  "lookup": ""},
            {"name": "created_at",  "type": "timestamptz","constraints": "NOT NULL DEFAULT now()","description": "Timestamp of comment creation. Used for DESC ordering.",               "lookup": ""},
        ],
    },
    {
        "name": "comment_likes",
        "status": "EXISTING — Remote (Supabase)",
        "layer": "remote",
        "description": (
            "Existing Supabase table. Tracks which users have liked which comments. "
            "Unlike is implemented as a DELETE by (comment_id, user_id)."
        ),
        "columns": [
            {"name": "id",         "type": "uuid", "constraints": "PK",           "description": "UUID primary key.",                          "lookup": ""},
            {"name": "comment_id", "type": "uuid", "constraints": "FK · NOT NULL","description": "The liked comment.",                         "lookup": "comments.id"},
            {"name": "user_id",    "type": "uuid", "constraints": "FK · NOT NULL","description": "The user who liked the comment.",            "lookup": "auth.users.id"},
        ],
    },
]

# ── Enum reference sheets ─────────────────────────────────────────────────────
ENUMS = [
    {
        "name": "Enum: GameMode",
        "description": "Valid values for StoredGameSession.modeRawValue / game_sessions.mode",
        "values": [
            ("Dink Sinks",    "Timed mode. Players score by achieving consecutive dink streaks within 60 seconds."),
            ("Volley Wallies","Timed mode. Players count every clean volley (≥ 15 MPH) in the round."),
            ("The Real Deal", "Manual rally scoring. First player to reach 5 points wins."),
            ("Pickle Cup",    "All three modes played in sequence. One overall champion declared."),
        ],
    },
    {
        "name": "Enum: DominantArm",
        "description": "Valid values for PlayerProfile.dominantArmRawValue",
        "values": [
            ("Right",         "Player uses their right arm as their dominant paddle arm."),
            ("Left",          "Player uses their left arm as their dominant paddle arm."),
            ("Ambidextrous",  "Player can play equally with either arm."),
        ],
    },
    {
        "name": "Enum: SkillLevel",
        "description": "Valid values for PlayerProfile.skillLevelRawValue",
        "values": [
            ("Beginner",      "New to pickleball. Learning basic rules and strokes."),
            ("Intermediate",  "Comfortable with most shots. Playing recreational leagues."),
            ("Advanced",      "Consistent technique. Competing in club or regional events."),
            ("Tournament",    "Competing at tournament level. High consistency and strategy."),
        ],
    },
    {
        "name": "Enum: BadgeType",
        "description": "Valid values for badges.badge_type CHECK constraint",
        "values": [
            ("achievement",  "Earned by reaching a stat milestone (e.g. 100 total hits, 75% sweet spot)."),
            ("milestone",    "Earned by reaching a participation milestone (e.g. 10 sessions, 30 days played)."),
            ("tournament",   "Awarded from results imported via GPN tournament integration."),
            ("level",        "Automatically awarded when a player crosses a level_tiers.xp_threshold."),
        ],
    },
    {
        "name": "Enum: SyncOperation",
        "description": "Valid values for SyncQueueItem.operation",
        "values": [
            ("upsert_profile",     "Push PlayerProfile changes to user_profiles (includes total_xp update)."),
            ("save_session",       "Push a new StoredGameSession to game_sessions."),
            ("upsert_location",    "Push a new or updated SavedLocation to saved_locations."),
            ("award_badge",        "Write a newly earned badge to user_badges."),
            ("xp_events",          "Append one or more rows to xp_events, then PATCH user_profiles.total_xp."),
        ],
    },
]


# ── Workbook builder ──────────────────────────────────────────────────────────

def make_wb():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Index sheet ──────────────────────────────────────────────────────────
    ws_idx = wb.create_sheet("Index")
    ws_idx.sheet_view.showGridLines = False
    ws_idx.column_dimensions["A"].width = 30
    ws_idx.column_dimensions["B"].width = 18
    ws_idx.column_dimensions["C"].width = 14
    ws_idx.column_dimensions["D"].width = 55

    # Title
    ws_idx.merge_cells("A1:D1")
    c = ws_idx["A1"]
    c.value = "DinkLink — Data Model Index"
    c.font = Font(name="Calibri", bold=True, size=16, color=CLR_WHITE)
    c.fill = fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_idx.row_dimensions[1].height = 30

    ws_idx.merge_cells("A2:D2")
    c = ws_idx["A2"]
    c.value = "Generated: April 13, 2026  |  Option A + MVP (13 tables)  |  Each tab = one table or enum"
    c.font = FONT_SMALL
    c.fill = fill("F0F4F8")
    c.alignment = Alignment(horizontal="center")

    headers = ["Table / Sheet", "Status", "Layer", "Description"]
    for col, h in enumerate(headers, 1):
        c = ws_idx.cell(row=4, column=col, value=h)
        c.font = FONT_COL_H
        c.fill = fill(CLR_HEADER_MID)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws_idx.row_dimensions[4].height = 20

    row = 5
    for t in TABLES:
        status_short = "Existing" if "EXISTING" in t["status"] else "New"
        layer_short = "Local" if t["layer"] == "local" else "Remote"
        row_fill = fill(CLR_EXISTING if status_short == "Existing" else CLR_NEW)
        for col, val in enumerate([t["name"], status_short, layer_short, t["description"]], 1):
            c = ws_idx.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.fill = row_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
        ws_idx.row_dimensions[row].height = 40
        row += 1

    # Enum rows
    for e in ENUMS:
        for col, val in enumerate([e["name"], "Reference", "—", e["description"]], 1):
            c = ws_idx.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.fill = fill("FFF8E1")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
        ws_idx.row_dimensions[row].height = 30
        row += 1

    # ── One sheet per table ──────────────────────────────────────────────────
    COL_WIDTHS = [28, 18, 30, 48, 40]
    COL_HEADERS = ["Column Name", "Data Type", "Constraints", "Description", "Linked To"]

    def safe_sheet_name(name):
        """Strip chars invalid in Excel sheet names and truncate to 31."""
        invalid = r'\/*?:[]'
        for ch in invalid:
            name = name.replace(ch, "_")
        return name[:31]

    for t in TABLES:
        ws = wb.create_sheet(safe_sheet_name(t["name"]))
        ws.sheet_view.showGridLines = False

        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Table title row
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = t["name"]
        c.font = Font(name="Calibri", bold=True, size=14, color=CLR_WHITE)
        c.fill = fill(CLR_HEADER_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 28

        # Status row
        ws.merge_cells("A2:E2")
        c = ws["A2"]
        c.value = t["status"]
        status_color = CLR_EXISTING if "EXISTING" in t["status"] else CLR_NEW
        c.font = Font(name="Calibri", bold=True, size=10, color="444444")
        c.fill = fill(status_color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

        # Description row
        ws.merge_cells("A3:E3")
        c = ws["A3"]
        c.value = t["description"]
        c.font = FONT_SMALL
        c.fill = fill("F7F9FC")
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[3].height = 45

        # Column headers
        for col, h in enumerate(COL_HEADERS, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = FONT_COL_H
            c.fill = fill(CLR_HEADER_MID)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[5].height = 20

        # Data rows
        for r_idx, col_def in enumerate(t["columns"], 6):
            is_pk = "PK" in col_def["constraints"]
            is_fk = "FK" in col_def["constraints"]
            if is_pk and is_fk:
                row_color = "FDE8B0"  # PK+FK: orange-ish
            elif is_pk:
                row_color = CLR_PK
            elif is_fk:
                row_color = CLR_FK
            else:
                row_color = CLR_WHITE if r_idx % 2 == 0 else CLR_HEADER_LIGHT

            values = [
                col_def["name"],
                col_def["type"],
                col_def["constraints"],
                col_def["description"],
                col_def["lookup"],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=r_idx, column=col, value=val)
                c.font = FONT_BODY_B if col == 1 else FONT_BODY
                c.fill = fill(row_color)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = BORDER
            ws.row_dimensions[r_idx].height = 36

        # Legend
        legend_row = len(t["columns"]) + 8
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Calibri", bold=True, size=9)
        legend_items = [
            (CLR_PK,   "PK — Primary Key"),
            (CLR_FK,   "FK — Foreign Key"),
            ("FDE8B0",  "PK + FK"),
            (CLR_HEADER_LIGHT, "Regular column (alternating)"),
        ]
        for i, (color, label) in enumerate(legend_items, 1):
            c = ws.cell(row=legend_row, column=i+1, value=label)
            c.fill = fill(color)
            c.font = FONT_SMALL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")

    # ── Enum reference sheets ────────────────────────────────────────────────
    for e in ENUMS:
        ws = wb.create_sheet(safe_sheet_name(e["name"]))
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 70

        ws.merge_cells("A1:B1")
        c = ws["A1"]
        c.value = e["name"]
        c.font = Font(name="Calibri", bold=True, size=13, color=CLR_WHITE)
        c.fill = fill("7B5EA7")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:B2")
        c = ws["A2"]
        c.value = e["description"]
        c.font = FONT_SMALL
        c.fill = fill("F3E5F5")
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[2].height = 30

        for col, h in enumerate(["Value", "Description"], 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = FONT_COL_H
            c.fill = fill("7B5EA7")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

        for r_idx, (val, desc) in enumerate(e["values"], 5):
            row_color = CLR_WHITE if r_idx % 2 == 0 else "F3E5F5"
            for col, cell_val in enumerate([val, desc], 1):
                c = ws.cell(row=r_idx, column=col, value=cell_val)
                c.font = FONT_BODY_B if col == 1 else FONT_BODY
                c.fill = fill(row_color)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = BORDER
            ws.row_dimensions[r_idx].height = 28

    return wb


if __name__ == "__main__":
    output_path = "/Users/julianeyman/DinkLink/DinkLink-DataModel.xlsx"
    wb = make_wb()
    wb.save(output_path)
    print(f"Saved: {output_path}")
